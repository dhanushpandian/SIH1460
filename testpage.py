import random
import PyPDF2
import openpyxl

# Define subjects and their corresponding file paths
subjects = {
    "Math": {
        "hard_pdf": "math_hard_questions.pdf",
        "easy_pdf": "math_easy_questions.pdf",
        "hard_excel": "math_hard_questions.xlsx",
        "easy_excel": "math_easy_questions.xlsx",
    },
    # Add more subjects here
}

def load_pdf_questions(pdf_path):
    questions = []
    pdf_reader = PyPDF2.PdfFileReader(open(pdf_path, "rb"))
    
    for page_num in range(pdf_reader.numPages):
        page = pdf_reader.getPage(page_num)
        text = page.extractText()
        questions.extend(text.split("\n"))
    
    return [question.strip() for question in questions if question.strip()]

def load_excel_questions(excel_path):
    questions = []
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        question = row[0]
        questions.append(question)
    
    return questions

def generate_test(subject):
    hard_questions_pdf = load_pdf_questions(subject["hard_pdf"])
    easy_questions_pdf = load_pdf_questions(subject["easy_pdf"])
    hard_questions_excel = load_excel_questions(subject["hard_excel"])
    easy_questions_excel = load_excel_questions(subject["easy_excel"])
    
    while True:
        question_type = random.choice(["hard", "easy"])
        if question_type == "hard":
            question = random.choice(hard_questions_pdf + hard_questions_excel)
        else:
            question = random.choice(easy_questions_pdf + easy_questions_excel)
        
        answer = input(f"Question ({question_type}): {question}\nYour answer: ").strip().lower()
        
        # Simulate grading here
        is_correct = random.choice([True, False])
        
        if is_correct:
            print("Correct!\n")
            if question_type == "hard":
                easy_questions_pdf.append(question)
                easy_questions_excel.append(question)
            else:
                hard_questions_pdf.append(question)
                hard_questions_excel.append(question)
        else:
            print("Incorrect!\n")
            if question_type == "hard":
                hard_questions_pdf.remove(question)
                hard_questions_excel.remove(question)
            else:
                easy_questions_pdf.remove(question)
                easy_questions_excel.remove(question)


if __name__ == "__main__":
    # Choose a subject (e.g., "Math") and call the generate_test function
    chosen_subject = "Math"
    if chosen_subject in subjects:
        generate_test(subjects[chosen_subject])
    else:
        print("Subject not found.")
